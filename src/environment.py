import json
from typing import Union
import openpyxl

from datatypes import Aircraft
class Airport():
    def __init__(self,airport_file):
        self._load_airport_data(airport_file)
    def _load_airport_data(self,airport_file):
        with open(airport_file,"r") as fs:
            data = json.load(fs)
            node_map = data["nodes"]
            # updated_map = deepcopy(node_map)
            # for i in node_map.keys():
            #     for j in node_map[i]:
            #         if j not in updated_map.keys():
            #             updated_map[j] = [i]
            #         elif i not in updated_map[j]:
            #             updated_map[j].append(i)
            self.dept_runways= data["dept_runways"]
            self.arrival_runways = data["arrival_runways"]
            self.gates = data["gates"]
            self.nodes = node_map
 

    def populate_waiting_dict(self)->dict[str,Union[Aircraft|None]]:
        carry = {}
        for i in self.arrival_runways:
            carry[i] = None
        for i in self.dept_runways:
            carry[i] = None
        for i in self.gates:
            carry[i] = None
        return carry

if __name__ == "__main__":
    with open("baseline_airport.json","r") as fs:
        data = json.load(fs)
    xlsx_data = openpyxl.load_workbook("nodes.xlsx")
    updated_nodes = {}
    for i in data["nodes"].keys():
        updated_nodes[i] = {"edges":data["nodes"][i],"x_pos":xlsx_data.active.cell(int(i)+1,2).value,"y_pos":xlsx_data.active.cell(int(i)+1,3).value}
    data["nodes"] = updated_nodes
    with open("baseline_airport.json","w") as fs:
        json.dump(data,fs)